"""
DDQ Automation — Document Ingestion
Reads PDF, DOCX, XLSX, TXT from a local folder and returns text chunks.
"""

import os
import re
from pathlib import Path
from typing import List, Dict

import pdfplumber
import docx2txt
from openpyxl import load_workbook


# ── Data structures ───────────────────────────────────────────────────────────

class Chunk:
    def __init__(self, source: str, page: int, text: str):
        self.source = source   # filename
        self.page   = page     # page / sheet number (1-based)
        self.text   = text.strip()

    def __repr__(self):
        return f"Chunk({self.source!r}, p{self.page}, {len(self.text)} chars)"


# ── Per-format extractors ─────────────────────────────────────────────────────

def _extract_pdf(path: Path) -> List[Dict]:
    pages = []
    try:
        with pdfplumber.open(path) as pdf:
            for i, page in enumerate(pdf.pages, 1):
                text = page.extract_text() or ""
                if text.strip():
                    pages.append({"page": i, "text": text})
    except Exception as e:
        print(f"  [WARN] PDF extraction failed for {path.name}: {e}")
    return pages


def _extract_docx(path: Path) -> List[Dict]:
    try:
        text = docx2txt.process(str(path))
        return [{"page": 1, "text": text}] if text.strip() else []
    except Exception as e:
        print(f"  [WARN] DOCX extraction failed for {path.name}: {e}")
        return []


def _extract_xlsx(path: Path) -> List[Dict]:
    pages = []
    try:
        wb = load_workbook(str(path), read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) for c in row if c is not None]
                if cells:
                    rows.append(" | ".join(cells))
            if rows:
                pages.append({"page": 1, "text": f"[Sheet: {sheet_name}]\n" + "\n".join(rows)})
        wb.close()
    except Exception as e:
        print(f"  [WARN] XLSX extraction failed for {path.name}: {e}")
    return pages


def _extract_txt(path: Path) -> List[Dict]:
    try:
        text = path.read_text(encoding="utf-8", errors="ignore")
        return [{"page": 1, "text": text}] if text.strip() else []
    except Exception as e:
        print(f"  [WARN] TXT extraction failed for {path.name}: {e}")
        return []


EXTRACTORS = {
    ".pdf":  _extract_pdf,
    ".docx": _extract_docx,
    ".doc":  _extract_docx,
    ".xlsx": _extract_xlsx,
    ".xls":  _extract_xlsx,
    ".txt":  _extract_txt,
    ".md":   _extract_txt,
    ".csv":  _extract_txt,
}


# ── Chunker ───────────────────────────────────────────────────────────────────

def _split_into_chunks(source: str, page: int, text: str,
                        size: int = 1200, overlap: int = 200) -> List[Chunk]:
    """Split a page of text into overlapping chunks."""
    chunks = []
    start = 0
    while start < len(text):
        end = start + size
        chunk_text = text[start:end]
        if chunk_text.strip():
            chunks.append(Chunk(source, page, chunk_text))
        if end >= len(text):
            break
        start = end - overlap
    return chunks


# ── Public API ────────────────────────────────────────────────────────────────

def ingest_folder(folder_path: str,
                  chunk_size: int = 1200,
                  chunk_overlap: int = 200) -> List[Chunk]:
    """
    Walk a local folder, extract text from all supported files,
    return a flat list of Chunk objects.
    """
    folder = Path(folder_path)
    if not folder.exists():
        raise FileNotFoundError(f"Folder not found: {folder_path}")

    all_chunks: List[Chunk] = []
    files = sorted([f for f in folder.iterdir() if f.is_file()])

    print(f"\n[Ingestion] Reading {len(files)} files from {folder.name}/")

    for f in files:
        ext = f.suffix.lower()
        extractor = EXTRACTORS.get(ext)
        if not extractor:
            print(f"  [SKIP] {f.name} (unsupported type {ext})")
            continue

        print(f"  [READ] {f.name}")
        pages = extractor(f)
        for pg in pages:
            chunks = _split_into_chunks(
                f.name, pg["page"], pg["text"], chunk_size, chunk_overlap
            )
            all_chunks.extend(chunks)

    print(f"[Ingestion] {len(all_chunks)} chunks extracted from {len(files)} files\n")
    return all_chunks


def chunks_to_context(chunks: List[Chunk], max_chars: int = 12000) -> str:
    """
    Concatenate chunks into a single context string for Claude,
    with source labels. Truncates at max_chars.
    """
    parts = []
    total = 0
    for c in chunks:
        header = f"\n--- [{c.source}, p.{c.page}] ---\n"
        body   = c.text
        if total + len(header) + len(body) > max_chars:
            break
        parts.append(header + body)
        total += len(header) + len(body)
    return "\n".join(parts)


def search_chunks(chunks: List[Chunk], keywords: List[str],
                  max_results: int = 12) -> List[Chunk]:
    """
    Simple keyword search over chunks — returns the most relevant chunks
    for a given set of keywords (case-insensitive substring match).
    Ranked by number of keyword hits.
    """
    scored = []
    kws_lower = [k.lower() for k in keywords]
    for c in chunks:
        text_lower = c.text.lower()
        score = sum(kw in text_lower for kw in kws_lower)
        if score > 0:
            scored.append((score, c))
    scored.sort(key=lambda x: -x[0])
    return [c for _, c in scored[:max_results]]

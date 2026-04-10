"""
DDQ Automation — Excel Underwriting Model Seed
Writes extracted deal metrics into a structured Excel workbook
pre-populated with formulas for DSCR, LCOE, BTC scenarios, and capex.
"""

import os
from datetime import datetime
from typing import Dict, Any

from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment,
                              Border, Side, numbers)
from openpyxl.utils import get_column_letter

from config import OUTPUT_DIR, COLOUR_NAVY, COLOUR_STEEL, COLOUR_LGRAY, COLOUR_GOLD


# ── Style helpers ─────────────────────────────────────────────────────────────

def _hex_fill(hex_colour: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_colour.upper())

def _border(style="thin"):
    s = Side(style=style, color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _header_font(size=10, bold=True, colour="FFFFFF"):
    return Font(name="Arial", size=size, bold=bold, color=colour)

def _body_font(size=9, bold=False, colour="1A1A1A"):
    return Font(name="Arial", size=size, bold=bold, color=colour)

def _label_font(size=9, bold=True, colour=COLOUR_NAVY):
    return Font(name="Arial", size=size, bold=bold, color=colour)

def _set_row(ws, row: int, values: list, fill=None, font=None,
             number_fmt: str = None, alignment=None):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        if fill:       cell.fill      = fill
        if font:       cell.font      = font
        if number_fmt: cell.number_format = number_fmt
        if alignment:  cell.alignment = alignment
        cell.border = _border()

def _col_widths(ws, widths: Dict[int, int]):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

def _section_header(ws, row: int, text: str, ncols: int = 5):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1, value=f"  {text}")
    cell.fill      = _hex_fill(COLOUR_NAVY)
    cell.font      = _header_font(10)
    cell.alignment = Alignment(vertical="center")
    cell.border    = _border()

def _input_row(ws, row: int, label: str, value, unit: str = "",
               note: str = "", number_fmt: str = None, input_col: int = 3):
    ws.cell(row=row, column=1, value=label).font    = _label_font()
    ws.cell(row=row, column=1).border = _border()
    ws.cell(row=row, column=2, value=unit).font     = _body_font(colour="5A6472")
    ws.cell(row=row, column=2).border = _border()

    cell = ws.cell(row=row, column=input_col, value=value)
    cell.font      = _body_font(bold=True)
    cell.fill      = _hex_fill("EFF7FF")
    cell.border    = _border()
    if number_fmt:
        cell.number_format = number_fmt

    if note:
        ws.cell(row=row, column=4, value=note).font = _body_font(colour="5A6472")
        ws.cell(row=row, column=4).border = _border()

def _formula_row(ws, row: int, label: str, formula: str, unit: str = "",
                 number_fmt: str = None):
    ws.cell(row=row, column=1, value=label).font    = _body_font(bold=True, colour=COLOUR_STEEL)
    ws.cell(row=row, column=1).border = _border()
    ws.cell(row=row, column=2, value=unit).font     = _body_font(colour="5A6472")
    ws.cell(row=row, column=2).border = _border()

    cell = ws.cell(row=row, column=3, value=formula)
    cell.font   = _body_font()
    cell.fill   = _hex_fill(COLOUR_LGRAY)
    cell.border = _border()
    if number_fmt:
        cell.number_format = number_fmt


# ── Sheet builders ────────────────────────────────────────────────────────────

def _build_deal_summary(wb: Workbook, schema: Dict[str, Any]):
    ws = wb.create_sheet("Deal Summary")
    ws.sheet_view.showGridLines = False
    _col_widths(ws, {1:32, 2:12, 3:18, 4:30})

    # Title
    ws.merge_cells("A1:D1")
    ws["A1"] = f"  {schema.get('deal_name','Project XXX')} — Deal Summary"
    ws["A1"].font      = Font(name="Arial", size=14, bold=True, color=COLOUR_NAVY)
    ws["A1"].fill      = _hex_fill(COLOUR_NAVY)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 28

    ws["A2"] = f"  Generated: {datetime.now().strftime('%d %b %Y %H:%M')}  |  BTSF DDQ Automation Tool"
    ws["A2"].font = Font(name="Arial", size=8, italic=True, color="5A6472")
    ws.merge_cells("A2:D2")

    row = 4
    sections = [
        ("SITE", [
            ("Site address",          schema.get("site_address"),         "",       ""),
            ("State",                 schema.get("state"),                "",       ""),
            ("County",                schema.get("county"),               "",       ""),
            ("Land area",             schema.get("land_area_acres"),       "acres",  "#,##0.0"),
            ("Site type",             schema.get("site_type"),             "",       ""),
            ("Lease / tenure",        schema.get("lease_term_years"),      "years",  "0"),
        ]),
        ("POWER", [
            ("Phase 1 capacity",      schema.get("capacity_mw_phase1"),    "MW",     "0.0"),
            ("Total potential",       schema.get("capacity_mw_total"),     "MW",     "0.0"),
            ("Energy source",         schema.get("energy_source"),         "",       ""),
            ("TDSP / utility",        schema.get("utility_tdsp"),          "",       ""),
            ("REP",                   schema.get("rep_name"),              "",       ""),
            ("All-in energy cost",    schema.get("energy_cost_mwh"),       "$/MWh",  "0.00"),
            ("Fixed PPA",             schema.get("ppa_present"),           "",       ""),
            ("RLO / energisation deadline", schema.get("fea_rlo_deadline"),"",       ""),
        ]),
        ("EQUIPMENT", [
            ("ASIC model",            schema.get("asic_model"),            "",       ""),
            ("ASIC quantity",         schema.get("asic_quantity"),         "units",  "#,##0"),
            ("Total hashrate",        schema.get("hashrate_total_ph"),     "PH/s",   "0.00"),
            ("Fleet efficiency",      schema.get("efficiency_jth"),        "J/TH",   "0.0"),
            ("Container type",        schema.get("container_type"),        "",       ""),
        ]),
        ("FINANCIAL", [
            ("Acquisition cost",      schema.get("acquisition_cost_usd"),  "USD",    "$#,##0"),
            ("Total capex Phase 1",   schema.get("total_capex_usd"),       "USD",    "$#,##0"),
            ("Facility size",         schema.get("facility_size_usd"),     "USD",    "$#,##0"),
            ("Facility tenor",        schema.get("facility_tenor_months"), "months", "0"),
            ("Target BTC IRR",        schema.get("target_irr_pct"),        "%",      "0.0%"),
            ("DSCR (base case)",      schema.get("dscr_base_case"),        "x",      "0.00x"),
            ("BTC breakeven",         schema.get("btc_breakeven_usd"),     "USD",    "$#,##0"),
            ("Annual energy cost",    schema.get("energy_cost_annual_usd"),"USD",    "$#,##0"),
        ]),
        ("LEGAL", [
            ("Phase 1 ESA complete",  schema.get("phase1_esa_complete"),   "",       ""),
            ("ALTA survey current",   schema.get("alta_survey_complete"),  "",       ""),
            ("Permits filed",         schema.get("permits_filed"),         "",       ""),
            ("Environmental indemnitor", schema.get("env_indemnitor"),     "",       ""),
            ("Existing debt",         schema.get("existing_debt"),         "",       ""),
            ("Borrowing entity",      schema.get("borrowing_entity"),      "",       ""),
            ("Jurisdiction",          schema.get("jurisdiction"),          "",       ""),
        ]),
    ]

    for sec_title, fields in sections:
        _section_header(ws, row, sec_title, 4)
        row += 1
        for label, val, unit, fmt in fields:
            _input_row(ws, row, label, val, unit, number_fmt=fmt if fmt else None)
            ws.row_dimensions[row].height = 16
            row += 1
        row += 1


def _build_assumptions(wb: Workbook, schema: Dict[str, Any]):
    """Inputs & Assumptions sheet — analyst edits these cells."""
    ws = wb.create_sheet("Assumptions")
    ws.sheet_view.showGridLines = False
    _col_widths(ws, {1:36, 2:10, 3:16, 4:36})

    ws["A1"] = "  INPUTS & ASSUMPTIONS — edit blue cells"
    ws["A1"].font      = Font(name="Arial", size=12, bold=True, color=COLOUR_NAVY)
    ws["A1"].fill      = _hex_fill(COLOUR_NAVY)
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 24

    ws["A2"] = "  Blue = data room input   |   Grey = calculated"
    ws["A2"].font = Font(name="Arial", size=8, italic=True, color="5A6472")
    ws.merge_cells("A2:D2")

    row = 4
    groups = [
        ("SITE & POWER", [
            ("Phase 1 capacity",       schema.get("capacity_mw_phase1") or 65,   "MW",      "0.0",  "Approved FEA capacity"),
            ("Load factor",            0.95,                                       "%",       "0%",   "Effective utilisation"),
            ("All-in energy cost",     schema.get("energy_cost_mwh") or 38.0,    "$/MWh",   "0.00", "Base case"),
            ("Energy cost — bear",     45.0,                                       "$/MWh",   "0.00", "High price scenario"),
            ("Energy cost — bull",     28.0,                                       "$/MWh",   "0.00", "Low price / optimised"),
            ("Annual curtailment",     0.05,                                       "%",       "0%",   "% of hours curtailed"),
        ]),
        ("MINING ECONOMICS", [
            ("Network hashrate",       870,                                        "EH/s",    "#,##0","Current estimate"),
            ("Block reward",           3.125,                                       "BTC",     "0.000","Post-halving"),
            ("Fleet efficiency",       schema.get("efficiency_jth") or 17.0,      "J/TH",    "0.0",  ""),
            ("BTC price — bear",       50000,                                       "USD",     "$#,##0",""),
            ("BTC price — base",       75000,                                       "USD",     "$#,##0",""),
            ("BTC price — bull",       100000,                                      "USD",     "$#,##0",""),
            ("BTC pool fee",           0.02,                                        "%",       "0%",   "Mining pool commission"),
        ]),
        ("CAPEX", [
            ("Site acquisition",       schema.get("acquisition_cost_usd") or 14000000, "USD","$#,##0",""),
            ("ASIC fleet",             None,                                        "USD",     "$#,##0","To be confirmed"),
            ("Infrastructure / EPC",   None,                                        "USD",     "$#,##0","To be confirmed"),
            ("Containers",             None,                                        "USD",     "$#,##0",""),
            ("Working capital",        None,                                        "USD",     "$#,##0",""),
            ("Contingency (10%)",      None,                                        "USD",     "$#,##0","Auto-calculated in model"),
        ]),
        ("DEBT", [
            ("Facility size",          schema.get("facility_size_usd") or 50000000,"USD",     "$#,##0",""),
            ("Tenor",                  schema.get("facility_tenor_months") or 36,  "months",  "0",    ""),
            ("Interest rate",          0.10,                                        "%",       "0.0%", "BTC-denominated yield"),
            ("Debt service reserve",   2,                                           "quarters","0",    "Quarters held in reserve"),
        ]),
        ("OPEX", [
            ("Monthly energy cost",    None,                                        "USD",     "$#,##0","Calculated from capacity × cost"),
            ("Monthly staff",          75000,                                       "USD",     "$#,##0",""),
            ("Maintenance",            20000,                                       "USD",     "$#,##0","Transformer/switchgear SLA"),
            ("Site overhead",          40000,                                       "USD",     "$#,##0","Water, security, insurance"),
            ("REP pre-payment buffer", 1,                                           "months",  "0",    "Months of energy cost held"),
        ]),
    ]

    for sec_title, fields in groups:
        _section_header(ws, row, sec_title, 4)
        row += 1
        for label, val, unit, fmt, note in fields:
            _input_row(ws, row, label, val, unit, note=note,
                       number_fmt=fmt if fmt else None)
            ws.row_dimensions[row].height = 16
            row += 1
        row += 1


def _build_scenarios(wb: Workbook):
    """Revenue & DSCR scenario matrix."""
    ws = wb.create_sheet("Scenarios")
    ws.sheet_view.showGridLines = False
    _col_widths(ws, {1:32, 2:10, 3:16, 4:16, 5:16})

    ws["A1"] = "  REVENUE & DSCR SCENARIOS"
    ws["A1"].font  = Font(name="Arial", size=12, bold=True, color=COLOUR_NAVY)
    ws["A1"].fill  = _hex_fill(COLOUR_NAVY)
    ws.merge_cells("A1:E1")
    ws.row_dimensions[1].height = 24

    # Column headers
    headers = ["Metric", "Unit", "Bear", "Base", "Bull"]
    fills   = [COLOUR_NAVY, COLOUR_NAVY, "8B0000", COLOUR_STEEL, "1D6A3A"]
    row = 3
    for col, (h, fill) in enumerate(zip(headers, fills), 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font  = _header_font(9)
        c.fill  = _hex_fill(fill)
        c.border = _border()
        c.alignment = Alignment(horizontal="center")

    # Reference Assumptions sheet (A = row index lookups)
    # Using simple placeholder formulas — analyst wires these to Assumptions sheet
    rows = [
        ("BTC price",              "USD",    "=Assumptions!C16", "=Assumptions!C17", "=Assumptions!C18"),
        ("Network hashrate",       "EH/s",   "=Assumptions!C12", "=Assumptions!C12", "=Assumptions!C12"),
        ("Capacity (MW)",          "MW",     "=Assumptions!C5",  "=Assumptions!C5",  "=Assumptions!C5"),
        ("Load factor",            "%",      "=Assumptions!C6",  "=Assumptions!C6",  "=Assumptions!C6"),
        ("Energy cost",            "$/MWh",  "=Assumptions!C8",  "=Assumptions!C7",  "=Assumptions!C9"),
        ("Annual energy (MWh)",    "MWh",    "=C4*C5*8760",      "=D4*D5*8760",      "=E4*E5*8760"),
        ("Annual energy cost",     "USD",    "=C9*C8",           "=D9*D8",           "=E9*E8"),
        ("Annual hashrate (EH/s)", "EH/s",   "=C4*0.001*C5",     "=D4*0.001*D5",     "=E4*0.001*E5"),
        ("Annual BTC mined",       "BTC",    "=C11/C3*6*24*365*3.125*(1-Assumptions!C19)",
                                             "=D11/D3*6*24*365*3.125*(1-Assumptions!C19)",
                                             "=E11/E3*6*24*365*3.125*(1-Assumptions!C19)"),
        ("Annual revenue",         "USD",    "=C12*C4",          "=D12*D4",          "=E12*E4"),
        ("Gross profit",           "USD",    "=C13-C10",         "=D13-D10",         "=E13-E10"),
        ("Gross margin",           "%",      "=C14/C13",         "=D14/D13",         "=E14/E13"),
        ("Annual debt service",    "USD",    "=Assumptions!C26/Assumptions!C27*12",
                                             "=Assumptions!C26/Assumptions!C27*12",
                                             "=Assumptions!C26/Assumptions!C27*12"),
        ("DSCR",                   "x",      "=C14/C16",         "=D14/D16",         "=E14/E16"),
    ]

    fmts = ["$#,##0","#,##0","0.0","0%","$#,##0","#,##0","$#,##0",
            "0.000","0.0","$#,##0","$#,##0","0.0%","$#,##0","0.00x"]

    for i, (label, unit, bear, base, bull) in enumerate(rows, 1):
        r = row + i
        ws.row_dimensions[r].height = 16
        bg = COLOUR_LGRAY if i % 2 == 0 else "FFFFFF"

        bold_rows = {12, 13, 15, 16}  # revenue, gross profit, DSCR
        fnt = _body_font(bold=(i in bold_rows))

        for col, val in enumerate([label, unit, bear, base, bull], 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font   = fnt if col > 2 else _label_font()
            c.fill   = _hex_fill(bg)
            c.border = _border()
            if col > 2 and i-1 < len(fmts):
                c.number_format = fmts[i-1]
            if col > 2:
                c.alignment = Alignment(horizontal="center")


def _build_capex(wb: Workbook, schema: Dict[str, Any]):
    """Capex breakdown sheet."""
    ws = wb.create_sheet("Capex")
    ws.sheet_view.showGridLines = False
    _col_widths(ws, {1:40, 2:16, 3:16, 4:28})

    ws["A1"] = "  CAPEX BREAKDOWN — Phase 1"
    ws["A1"].font  = Font(name="Arial", size=12, bold=True, color=COLOUR_NAVY)
    ws["A1"].fill  = _hex_fill(COLOUR_NAVY)
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 24

    headers = ["Item", "Amount (USD)", "% of total", "Source / note"]
    row = 3
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font   = _header_font(9)
        c.fill   = _hex_fill(COLOUR_STEEL)
        c.border = _border()

    items = [
        ("Site acquisition",                    schema.get("acquisition_cost_usd"),  "MIPA"),
        ("ASIC mining equipment",               None,                                "TBC — market pricing at drawdown"),
        ("Mining containers",                   None,                                "TBC"),
        ("Transformer maintenance + arc flash", 200000,                              "Electrical contractor estimate"),
        ("Cooling infrastructure",              720000,                              "Estimate"),
        ("Electrical installation (cabling, PDUs)", 4068000,                         "Estimate"),
        ("Site civil works",                    125000,                              "Estimate"),
        ("Fiber + networking",                  132000,                              "Estimate"),
        ("Working capital + REP deposit",       None,                                "TBC — depends on Ammper TC"),
        ("Transaction costs + legal + permits", None,                                "TBC"),
        ("Contingency (10%)",                   None,                                "Auto-calculated"),
    ]

    data_start = row + 1
    for i, (label, amount, note) in enumerate(items, 1):
        r = row + i
        bg = COLOUR_LGRAY if i % 2 == 0 else "FFFFFF"
        ws.row_dimensions[r].height = 16

        ws.cell(r, 1, label).font   = _body_font()
        ws.cell(r, 1).fill          = _hex_fill(bg)
        ws.cell(r, 1).border        = _border()

        c_amt = ws.cell(r, 2, amount)
        c_amt.font   = _body_font(bold=True)
        c_amt.fill   = _hex_fill("EFF7FF" if amount is not None else bg)
        c_amt.border = _border()
        c_amt.number_format = "$#,##0"

        c_pct = ws.cell(r, 3)
        c_pct.font   = _body_font()
        c_pct.fill   = _hex_fill(COLOUR_LGRAY)
        c_pct.border = _border()
        c_pct.number_format = "0.0%"

        ws.cell(r, 4, note).font   = _body_font(colour="5A6472")
        ws.cell(r, 4).border       = _border()

    # Total row
    total_row = row + len(items) + 1
    ws.row_dimensions[total_row].height = 18
    ws.cell(total_row, 1, "TOTAL CAPEX").font   = _label_font()
    ws.cell(total_row, 1).fill                  = _hex_fill(COLOUR_NAVY)
    ws.cell(total_row, 1).border                = _border()
    c_tot = ws.cell(total_row, 2, f"=SUM(B{data_start}:B{total_row-1})")
    c_tot.font          = _header_font(10)
    c_tot.fill          = _hex_fill(COLOUR_NAVY)
    c_tot.border        = _border()
    c_tot.number_format = "$#,##0"


# ── Main export function ──────────────────────────────────────────────────────

def export_excel(schema: Dict[str, Any], deal_name: str = "Project XXX") -> str:
    """
    Generate the Excel underwriting model seed.
    Returns the path to the saved file.
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    wb = Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    print("[Excel] Building underwriting model...")
    _build_deal_summary(wb, schema)
    _build_assumptions(wb, schema)
    _build_scenarios(wb)
    _build_capex(wb, schema)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    safe_name = deal_name.replace(" ", "_").replace("/", "-")
    filename  = f"Underwriting_{safe_name}_{timestamp}.xlsx"
    filepath  = os.path.join(OUTPUT_DIR, filename)
    wb.save(filepath)

    print(f"[Excel] Saved: {filepath}")
    return filepath
